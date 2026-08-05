# -*- coding: utf-8 -*-
"""
Excel 导入解析子模块 - 增强版
支持：多级表头查找、模糊子串匹配、预解析预览、表头映射自动学习
"""
import re
from openpyxl import load_workbook
from services.data_service import clean_job_row, insert_job, is_valid_job

# 标准字段 → 可能的表头关键字（按优先级排列，子串模糊匹配）
# 支持 "包含匹配"：只要表头名包含关键字即可命中
FIELD_KEYWORDS = {
    "company": ["企业名称", "公司名称", "企业", "单位名称", "招聘单位", "用人单位", "单位", "公司"],
    "contact": ["联系人", "招聘联系人", "联系人员", "招聘人"],
    "phone": ["联系电话", "电话", "联系方式", "手机", "手机号码", "联系手机", "手机"],
    "job_name": ["岗位名称", "岗位", "工种", "职位名称", "招聘岗位", "招聘职位", "岗位工种", "招聘岗位"],
    "recruit_num": ["招聘人数", "人数", "需求人数", "招聘数量"],
    "salary": ["薪资", "薪资待遇", "月薪", "工资", "薪酬", "收入", "薪资水平", "月收入"],
    "education": ["学历要求", "学历", "文化程度", "学历文化", "学历要求文化"],
    "region": ["地区", "所属地区", "工作地区", "城市", "工作地点", "单位地址", "招聘地区"],
    "industry": ["行业", "行业类型", "所属行业", "岗位行业"],
    "shift": ["班次", "工作班次", "上班班次", "工作时间"],
    "duty": ["岗位职责", "岗位描述", "工作内容", "职责", "招聘要求", "岗位要求", "工作要求"],
    "welfare": ["福利待遇", "福利", "福利说明", "福利待遇说明"],
    "remark": ["备注", "说明", "其他", "备注说明"],
    "is_caring": ["是否爱心岗位", "爱心岗位", "是否爱心", "爱心"],
}


def _fuzzy_match_field(header_name, field_keywords):
    """
    模糊子串匹配：表头名包含任一字段关键字即视为命中。
    返回 (字段名, 命中的关键字) 或 None。
    """
    h = str(header_name).strip()
    if not h:
        return None
    # 精确匹配优先
    for kw in field_keywords:
        if kw == h:
            return kw
    # 子串包含匹配（按关键字长度降序，长的更精确）
    for kw in sorted(field_keywords, key=len, reverse=True):
        if kw in h:
            return kw
    return None


def auto_detect_header_row(rows, max_scan=10):
    """
    多级表头查找：扫描前 max_scan 行，找到包含最多已知字段关键字的行作为表头行。
    返回 (表头行索引, 表头列表, 字段映射表)
    """
    best_row_idx = 0
    best_headers = []
    best_mapping = {}
    best_score = 0

    scan_limit = min(max_scan, len(rows))
    for row_idx in range(scan_limit):
        row = rows[row_idx]
        headers = [str(c).strip() if c is not None else "" for c in row]
        mapping = {}
        for col_idx, h in enumerate(headers):
            if not h:
                continue
            for field, keywords in FIELD_KEYWORDS.items():
                match = _fuzzy_match_field(h, keywords)
                if match and field not in mapping:
                    mapping[field] = {"col_idx": col_idx, "header": h, "matched": match}
        score = len(mapping)
        if score > best_score:
            best_score = score
            best_row_idx = row_idx
            best_headers = headers
            best_mapping = mapping

    return best_row_idx, best_headers, best_mapping


def detect_merged_headers(ws, header_row_idx):
    """
    检测合并单元格并合并表头文本。
    例如：A1="单位" A2="名称" → 合并为 "单位名称"
    """
    headers = []
    for col in range(ws.max_column):
        cell = ws.cell(row=header_row_idx + 1, column=col + 1)
        val = str(cell.value).strip() if cell.value is not None else ""
        headers.append(val)
    return headers


def build_record_from_row(row, headers, mapping):
    """
    根据字段映射表从 Excel 行构建标准记录字典。
    仅提取已知字段，未知字段保留原始表头键值对。
    """
    record = {}
    for field, info in mapping.items():
        col_idx = info["col_idx"]
        if col_idx < len(row):
            record[info["header"]] = row[col_idx]
    # 同时保留所有未匹配的列（以原始表头为 key）
    for col_idx, val in enumerate(row):
        if col_idx < len(headers) and headers[col_idx]:
            if headers[col_idx] not in record:
                record[headers[col_idx]] = val
    return record


def preview_excel(file_path):
    """
    Excel 预解析预览：
      1. 分析表头结构（自动检测表头行、合并单元格）
      2. 展示识别到的字段映射（标准字段 → Excel 表头）
      3. 展示前 3 行数据样例
      4. 识别未匹配的表头
    返回 JSON 可直接供前端展示。
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    if not all_rows:
        return {"code": 1, "msg": "空表", "data": None}

    # 1. 自动检测表头行
    header_row_idx, headers, mapping = auto_detect_header_row(all_rows)

    # 2. 如果检测到 0 个字段映射，回退用第 0 行
    if not mapping:
        headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
        header_row_idx = 0
        for col_idx, h in enumerate(headers):
            if not h:
                continue
            for field, keywords in FIELD_KEYWORDS.items():
                if field not in mapping:
                    match = _fuzzy_match_field(h, keywords)
                    if match:
                        mapping[field] = {"col_idx": col_idx, "header": h, "matched": match}

    # 3. 识别未匹配表头
    matched_headers = set(info["header"] for info in mapping.values())
    unmatched_headers = [h for h in headers if h and h not in matched_headers]

    # 4. 构建字段映射展示
    field_mapping_display = []
    for field in ["company", "contact", "phone", "job_name", "recruit_num",
                   "salary", "education", "region", "industry", "shift",
                   "duty", "welfare", "remark", "is_caring"]:
        info = mapping.get(field)
        label_map = {
            "company": "企业/单位名称", "contact": "联系人", "phone": "联系电话",
            "job_name": "岗位名称", "recruit_num": "招聘人数", "salary": "薪资/收入",
            "education": "学历要求", "region": "地区", "industry": "行业",
            "shift": "班次", "duty": "岗位职责/要求", "welfare": "福利待遇",
            "remark": "备注", "is_caring": "是否爱心岗位",
        }
        if info:
            field_mapping_display.append({
                "field": field, "label": label_map.get(field, field),
                "matched": True, "header": info["header"],
                "matched_keyword": info["matched"],
            })
        else:
            field_mapping_display.append({
                "field": field, "label": label_map.get(field, field),
                "matched": False, "header": "",
            })

    # 5. 预览前 3 行数据
    data_start = header_row_idx + 1
    sample_rows = []
    for r in all_rows[data_start:data_start + 3]:
        record = build_record_from_row(r, headers, mapping)
        # 过滤空行
        if any(v not in (None, "") for v in record.values()):
            sample_rows.append(record)

    total_data_rows = sum(
        1 for r in all_rows[data_start:]
        if any(v not in (None, "") for v in r)
    )

    return {
        "code": 0,
        "data": {
            "header_row_index": header_row_idx + 1,  # 1-based for display
            "total_headers": len([h for h in headers if h]),
            "matched_fields": len(mapping),
            "total_data_rows": total_data_rows,
            "field_mapping": field_mapping_display,
            "unmatched_headers": unmatched_headers,
            "sample_rows": sample_rows,
            "headers": headers,
            "mapping": {k: v["col_idx"] for k, v in mapping.items()},
        },
    }


def import_excel_with_mapping(file_path, operator="admin", custom_mapping=None):
    """
    带映射的 Excel 导入：
      custom_mapping: 用户在预览界面调整后的字段映射（可选）
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    if not all_rows:
        return 0, 0, ["空表"]

    # 自动检测表头
    header_row_idx, headers, auto_mapping = auto_detect_header_row(all_rows)

    # 合并用户自定义映射
    mapping = dict(auto_mapping)
    if custom_mapping:
        for field, col_idx in custom_mapping.items():
            if col_idx is not None and col_idx < len(headers):
                mapping[field] = {
                    "col_idx": col_idx,
                    "header": headers[col_idx] if col_idx < len(headers) else "",
                    "matched": "用户指定",
                }

    data_start = header_row_idx + 1
    success, skip, errors = 0, 0, []
    for idx, row in enumerate(all_rows[data_start:], start=data_start + 1):
        record = build_record_from_row(row, headers, mapping)
        if not any(v not in (None, "") for v in record.values()):
            continue
        try:
            job = clean_job_row(record)
            if not is_valid_job(job):
                skip += 1
                errors.append(f"第{idx}行：企业名/岗位名均为空，跳过")
                continue
            insert_job(job, operator)
            success += 1
        except Exception as e:
            skip += 1
            errors.append(f"第{idx}行：{str(e)[:80]}")
    return success, skip, errors

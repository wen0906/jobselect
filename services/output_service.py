# -*- coding: utf-8 -*-
"""
模块四：输出与详情生成模块
子功能：标准化 Excel 报表导出、单岗位静态详情 HTML 生成、打印样式适配
核心自研逻辑：劳务对接报表模板渲染、独立岗位详情页面自动生成代码。
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import models
from services.filter_service import _build_where
from services.data_service import get_job_labels
from config import DETAIL_PAGE_DIR, EXPORT_DIR


# ============ 1. 标准化劳务对接 Excel 报表导出 ============
def export_jobs_excel(filters, filename=None):
    """
    自主编码实现标准化劳务对接 Excel 报表导出。
    返回生成的文件绝对路径。
    """
    where_sql, params = _build_where(filters)
    sql = (
        f"SELECT j.* FROM job j WHERE {where_sql} "
        f"ORDER BY j.is_caring DESC, j.max_salary DESC, j.id DESC"
    )
    rows = models.query_all(sql, params)

    wb = Workbook()
    ws = wb.active
    ws.title = "劳务对接岗位报表"

    # 标题行
    title = "东西部劳务协作岗位对接报表"
    ws.merge_cells("A1:K1")
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(name="黑体", size=16, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 副标题：导出时间 + 条数
    ws.merge_cells("A2:K2")
    sub = ws["A2"]
    sub.value = f"导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}    共 {len(rows)} 条岗位"
    sub.font = Font(name="宋体", size=10, italic=True, color="666666")
    sub.alignment = Alignment(horizontal="right")
    ws.row_dimensions[2].height = 18

    # 表头
    headers = ["序号", "地区", "企业名称", "岗位名称", "招聘人数",
               "薪资待遇", "学历要求", "行业", "班次", "福利待遇", "联系电话"]
    header_row = 3
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(name="黑体", size=11, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor="305496")
    ws.row_dimensions[header_row].height = 24

    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, r in enumerate(rows, start=1):
        labels = get_job_labels(r["id"])
        row_idx = header_row + i
        values = [
            i, r["region"], r["company"], r["job_name"], r["recruit_num"],
            r["salary_text"], r["education"], r["industry"], r["shift"],
            "、".join(labels), r["phone"],
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
            c.font = Font(name="宋体", size=10)
            # 爱心岗位整行高亮
            if r["is_caring"]:
                c.fill = PatternFill("solid", fgColor="FFF2CC")
        ws.row_dimensions[row_idx].height = 28

    # 列宽
    widths = [6, 8, 28, 18, 8, 14, 10, 12, 10, 30, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # 备注
    note_row = header_row + len(rows) + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=11)
    nc = ws.cell(row=note_row, column=1)
    nc.value = "说明：黄色底纹标记为爱心帮扶岗位，优先推荐大龄、低学历、无技能困难群体。"
    nc.font = Font(name="宋体", size=9, italic=True, color="C00000")

    if not filename:
        filename = f"劳务对接报表_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    path = os.path.join(EXPORT_DIR, filename)
    wb.save(path)
    return path


def export_contacts_excel(filters):
    """批量导出企业联系人、联系电话，用于线下招工通知。"""
    where_sql, params = _build_where(filters)
    sql = (
        f"SELECT j.company,j.contact,j.phone,j.region,j.job_name "
        f"FROM job j WHERE {where_sql} ORDER BY j.region,j.company"
    )
    rows = models.query_all(sql, params)
    wb = Workbook()
    ws = wb.active
    ws.title = "企业联系人"
    headers = ["地区", "企业名称", "联系人", "联系电话", "招聘岗位"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(name="黑体", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
        c.alignment = Alignment(horizontal="center")
    for i, r in enumerate(rows, start=2):
        for col, k in enumerate(["region", "company", "contact", "phone", "job_name"], start=1):
            ws.cell(row=i, column=col, value=r[k])
    for i, w in enumerate([10, 28, 12, 16, 20], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    path = os.path.join(EXPORT_DIR, f"企业联系人_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx")
    wb.save(path)
    return path


# ============ 2. 单岗位独立静态详情页生成（自研页面生成代码） ============
DETAIL_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=1,user-scalable=no">
<title>{company} - {job_name} 招聘详情</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:"PingFang SC","Microsoft YaHei",sans-serif;background:#f5f7fa;color:#2d3748;padding:16px}}
.card{{max-width:680px;margin:0 auto;background:#fff;border-radius:12px;box-shadow:0 6px 20px rgba(0,0,0,.08);overflow:hidden}}
.head{{background:linear-gradient(135deg,#3b82f6,#10b981);color:#fff;padding:24px 20px}}
.head .region{{display:inline-block;background:rgba(255,255,255,.25);padding:2px 10px;border-radius:12px;font-size:12px;margin-bottom:8px}}
.head h1{{font-size:22px;margin-bottom:6px}}
.head .company{{font-size:14px;opacity:.95}}
.caring{{display:inline-block;background:#fbbf24;color:#7c2d12;font-weight:bold;padding:3px 10px;border-radius:6px;font-size:12px;margin-left:8px}}
.body{{padding:20px}}
.salary{{font-size:26px;color:#ef4444;font-weight:bold;margin-bottom:4px}}
.salary small{{font-size:13px;color:#94a3b8;font-weight:normal}}
.row{{display:flex;padding:10px 0;border-bottom:1px dashed #e2e8f0;font-size:14px}}
.row .k{{width:90px;color:#94a3b8;flex-shrink:0}}
.row .v{{flex:1;color:#2d3748}}
.labels{{margin-top:14px;display:flex;flex-wrap:wrap;gap:8px}}
.labels .tag{{background:#ebf5ff;color:#3b82f6;padding:4px 10px;border-radius:6px;font-size:12px;border:1px solid #bfdbfe}}
.section{{margin-top:18px}}
.section h3{{font-size:15px;color:#3b82f6;border-left:4px solid #3b82f6;padding-left:8px;margin-bottom:10px}}
.section p{{font-size:14px;line-height:1.8;color:#4a5568;white-space:pre-wrap}}
.contact{{margin-top:20px;background:#f0f9ff;padding:16px;border-radius:8px;text-align:center}}
.contact .phone{{font-size:24px;color:#ef4444;font-weight:bold;margin:6px 0}}
.foot{{text-align:center;color:#94a3b8;font-size:12px;padding:14px}}
@media print{{body{{background:#fff;padding:0}}.card{{box-shadow:none;border-radius:0}}}}
</style>
</head>
<body>
<div class="card">
  <div class="head">
    <span class="region">{region}</span>{caring_badge}
    <h1>{job_name}</h1>
    <div class="company">{company}</div>
  </div>
  <div class="body">
    <div class="salary">{salary_text} <small>/ 月</small></div>
    <div class="row"><div class="k">招聘人数</div><div class="v">{recruit_num} 人</div></div>
    <div class="row"><div class="k">学历要求</div><div class="v">{education}</div></div>
    <div class="row"><div class="k">所属行业</div><div class="v">{industry}</div></div>
    <div class="row"><div class="k">工作班次</div><div class="v">{shift}</div></div>
    <div class="row"><div class="k">持证要求</div><div class="v">{cert_text}</div></div>
    <div class="labels">{labels_html}</div>
    <div class="section">
      <h3>岗位职责</h3>
      <p>{job_duty}</p>
    </div>
    <div class="section">
      <h3>福利待遇</h3>
      <p>{welfare_text}</p>
    </div>
    <div class="contact">
      <div>意向求职者请联系</div>
      <div class="phone">{phone}</div>
      <div>联系人：{contact}</div>
    </div>
  </div>
  <div class="foot">岗位智能筛选可视化管理系统 V1.0 · 生成于 {gen_time}</div>
</div>
</body>
</html>
"""


def generate_detail_page(job_id):
    """单岗位独立静态详情网页自动生成，适配手机转发、线下打印。"""
    job = models.query_one("SELECT * FROM job WHERE id=?", (job_id,))
    if not job:
        return None
    labels = get_job_labels(job_id)
    labels_html = "".join(f'<span class="tag">{l}</span>' for l in labels)
    caring_badge = '<span class="caring">爱心岗位</span>' if job["is_caring"] else ""
    cert_text = job["cert_text"] or ("持证优先" if job["has_cert_priority"] else "无要求")

    html = DETAIL_TEMPLATE.format(
        company=job["company"],
        job_name=job["job_name"],
        region=job["region"] or "—",
        caring_badge=caring_badge,
        salary_text=job["salary_text"],
        recruit_num=job["recruit_num"],
        education=job["education"],
        industry=job["industry"],
        shift=job["shift"],
        cert_text=cert_text,
        labels_html=labels_html,
        job_duty=job["job_duty"] or "—",
        welfare_text=job["welfare_text"] or "—",
        phone=job["phone"] or "—",
        contact=job["contact"] or "—",
        gen_time=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    # 生成独立 HTML 文件
    fname = f"job_{job_id}_{job['company'][:10]}.html"
    # 文件名安全处理
    safe_fname = "".join(c for c in fname if c not in r'\/:*?"<>|')
    path = os.path.join(DETAIL_PAGE_DIR, safe_fname)
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    return path
